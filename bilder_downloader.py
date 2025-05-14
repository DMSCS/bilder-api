import os
import re
import time
import base64
import requests
import mimetypes
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# === KONFIGURATION ===
BASE_URL = "https://www.skpk.de"
DOMAIN = urlparse(BASE_URL).netloc.replace("www.", "")
TIMESTAMP = datetime.now().strftime("%Y-%m-%d_%H-%M")
BASE_FOLDER_NAME = f"{DOMAIN}_{TIMESTAMP}"
BASE_DOWNLOAD_DIR = os.path.join("Bilder", BASE_FOLDER_NAME)
EXCEL_FILE = os.path.join(BASE_DOWNLOAD_DIR, "bilder_log.xlsx")
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

log_data = []

# === Selenium Setup ===
firefox_options = Options()
firefox_options.add_argument("--headless")
driver = webdriver.Firefox(options=firefox_options)

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

downloaded_files = set()

def make_soup(url):
    try:
        driver.get(url)
        time.sleep(3)
        return BeautifulSoup(driver.page_source, "html.parser")
    except Exception as e:
        print(f"[Fehler] Seite nicht ladbar: {url} -> {e}")
        return None

def get_navigation_links(soup):
    nav_links = {}
    nav = soup.find("nav")
    if nav:
        links = nav.find_all("a", href=True)
        for link in links:
            name = link.get_text(strip=True) or "Unbenannt"
            href = urljoin(BASE_URL, link["href"])
            nav_links[name] = href
    else:
        print("[Warnung] Kein <nav>-Element gefunden.")
    return nav_links

def is_allowed_filetype(filename):
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXTENSIONS

def log_to_excel(seite, bild_url, alt_text, pfad, typ):
    log_data.append({
        "Seite": seite,
        "Bild-URL": bild_url,
        "ALT-Text": alt_text,
        "Gespeichert als": pfad,
        "Typ": typ
    })

def download_image(img_url, folder_name, page_url, alt_text="", typ=""):
    try:
        folder_path = os.path.join(BASE_DOWNLOAD_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        if img_url.startswith("data:"):
            header, encoded = img_url.split(",", 1)
            mime_type = header.split(";")[0].replace("data:", "")
            extension = mimetypes.guess_extension(mime_type) or ".img"
            if extension not in ALLOWED_EXTENSIONS:
                return

            img_name = f"embedded_{abs(hash(img_url))}{extension}"
            img_path = os.path.join(folder_path, img_name)

            if img_path in downloaded_files:
                return

            with open(img_path, "wb") as f:
                f.write(base64.b64decode(encoded))

            downloaded_files.add(img_path)
            log_to_excel(page_url, img_url, alt_text, img_path, typ)

        else:
            img_name = os.path.basename(urlparse(img_url).path)
            if not img_name:
                img_name = "bild.jpg"

            if not is_allowed_filetype(img_name):
                return

            img_path = os.path.join(folder_path, img_name)
            if img_path in downloaded_files:
                return

            response = requests.get(img_url, headers=HEADERS, stream=True)
            response.raise_for_status()

            with open(img_path, "wb") as f:
                for chunk in response.iter_content(1024):
                    f.write(chunk)

            downloaded_files.add(img_path)
            log_to_excel(page_url, img_url, alt_text, img_path, typ)

    except Exception as e:
        print(f"[Fehler] Bild konnte nicht geladen werden: {img_url[:60]}... -> {e}")

def extract_background_images(page_url, folder_name):
    driver.get(page_url)
    time.sleep(3)
    folder_path = os.path.join(BASE_DOWNLOAD_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    elements = driver.find_elements(By.CSS_SELECTOR, "div")

    for element in elements:
        bg = element.value_of_css_property("background-image")
        match = re.search(r'url\("?([^")]+)"?\)', bg)
        if match:
            bg_url = match.group(1)
            if not bg_url.startswith("data:"):
                full_url = urljoin(page_url, bg_url)
                download_image(full_url, folder_name, page_url, alt_text="", typ="background")

def extract_and_download_images(page_url, folder_name):
    soup = make_soup(page_url)
    if soup:
        images = soup.find_all("img", src=True)
        for img in images:
            img_url = urljoin(page_url, img["src"])
            alt_text = img.get("alt", "").strip()
            download_image(img_url, folder_name, page_url, alt_text=alt_text, typ="<img>")

    extract_background_images(page_url, folder_name)

def write_excel_log(data, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilder-Log"

    headers = ["Seite", "Bild-URL", "ALT-Text", "Gespeichert als", "Typ"]
    ws.append(headers)

    for entry in data:
        row = [
            entry["Seite"],
            entry["Bild-URL"],
            entry["ALT-Text"],
            entry["Gespeichert als"],
            entry["Typ"]
        ]
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=2)
        url = entry["Bild-URL"]
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single")

    wb.save(filepath)

def main():
    os.makedirs(BASE_DOWNLOAD_DIR, exist_ok=True)

    homepage = make_soup(BASE_URL)
    if not homepage:
        print("[Abbruch] Startseite konnte nicht geladen werden.")
        return

    nav_links = get_navigation_links(homepage)

    if nav_links:
        for nav_name, nav_url in nav_links.items():
            print(f"\n[🔍] Verarbeite Navigationspunkt: '{nav_name}' -> {nav_url}")
            extract_and_download_images(nav_url, nav_name)
    else:
        print("[Info] Keine Navigation gefunden – analysiere stattdessen die Startseite selbst.")
        extract_and_download_images(BASE_URL, "Unkategorisiert")

    write_excel_log(log_data, EXCEL_FILE)
    print(f"\n✅ Fertig. Bilder und Log gespeichert in: {BASE_DOWNLOAD_DIR}")
    driver.quit()

if __name__ == "__main__":
    main()
